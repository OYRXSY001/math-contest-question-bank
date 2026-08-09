from collections import Counter
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from question_bank.katex import validate_markdown_formulas
from question_bank.models import (
    MAX_PDF_BYTES,
    KnowledgePoint,
    Paper,
    Question,
    QuestionKnowledgePoint,
)
from question_bank.templatetags.content import render_markdown

INVENTORY_HEADERS = [
    "edition", "stage", "original_category_label", "title", "exam_year",
    "source_url", "pdf_file", "question_count",
]
QUESTION_HEADERS = [
    "edition", "stage", "question_no", "sort_order", "question_type", "score",
    "primary_knowledge", "secondary_knowledge", "stem_md", "answer_md",
    "solution_md", "image_files", "source_page", "source_crop", "ocr_confidence",
    "text_checked", "formula_checked", "solution_checked", "unresolved_ocr_items",
    "katex_errors", "reviewer",
]
MIN_OCR_CONFIDENCE = Decimal("0.90")
MAX_IMAGE_BYTES = 10 * 1024 * 1024
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
SUSPICIOUS_OCR_CHARACTERS = {
    "�": "Unicode 替换字符",
    "□": "空方框",
}


def as_int(value, label, issues):
    if isinstance(value, bool):
        issues.append(f"{label}: 必须是整数")
        return 0
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        issues.append(f"{label}: 必须是整数")
        return 0
    if not isinstance(value, (str, bytes)) and value != parsed:
        issues.append(f"{label}: 必须是整数")
        return 0
    return parsed


def as_bool(value, label, issues):
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "是"}:
        return True
    if normalized in {"0", "false", "no", "否"}:
        return False
    issues.append(f"{label}: 必须是布尔值")
    return False


def as_ocr_confidence(value, label, issues):
    if value in (None, "") or isinstance(value, bool):
        issues.append(f"{label}: 必须是 0 到 1 之间的数字")
        return None
    try:
        confidence = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        issues.append(f"{label}: 必须是 0 到 1 之间的数字")
        return None
    if not confidence.is_finite() or not Decimal("0") <= confidence <= Decimal("1"):
        issues.append(f"{label}: 必须在 0 到 1 之间")
        return None
    return confidence


def validate_ocr_characters(row, label, issues):
    for field in ("stem_md", "answer_md", "solution_md"):
        text = str(row.get(field) or "")
        for character, name in SUSPICIOUS_OCR_CHARACTERS.items():
            if character in text:
                issues.append(f"{label} {field}: 包含{name} {character}")


def split_values(value):
    return [item.strip() for item in str(value or "").replace("，", ",").split(",") if item.strip()]


def safe_file(root, relative, label, issues, pdf=False):
    if not relative:
        return
    root = Path(root).resolve()
    target = (root / str(relative)).resolve()
    if not target.is_relative_to(root):
        issues.append(f"{label}: 路径超出允许目录")
        return
    if not target.is_file():
        issues.append(f"{label}: 文件不存在 {relative}")
        return
    if pdf:
        if target.suffix.lower() != ".pdf":
            issues.append(f"{label}: 文件扩展名必须是 .pdf")
        if target.stat().st_size > MAX_PDF_BYTES:
            issues.append(f"{label}: PDF 不能超过 50 MB")
        with target.open("rb") as source:
            if source.read(5) != b"%PDF-":
                issues.append(f"{label}: 文件内容不是 PDF")


def image_signature_matches(path):
    with path.open("rb") as source:
        header = source.read(12)
    suffix = path.suffix.lower()
    if suffix == ".png":
        return header.startswith(b"\x89PNG\r\n\x1a\n")
    if suffix in {".jpg", ".jpeg"}:
        return header.startswith(b"\xff\xd8\xff")
    if suffix == ".webp":
        return header.startswith(b"RIFF") and header[8:12] == b"WEBP"
    return False


class _ImageSourceParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.sources = set()

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "img":
            source = dict(attrs).get("src")
            if source is not None:
                self.sources.add(source)

    def handle_startendtag(self, tag, attrs):
        self.handle_starttag(tag, attrs)


def rendered_image_sources(markdown_sources):
    parser = _ImageSourceParser()
    for source in markdown_sources:
        parser.feed(str(render_markdown(source)))
    parser.close()
    return parser.sources


def validate_question_image(root, relative, label, markdown_sources, issues):
    root = Path(root).resolve()
    relative = str(relative)
    normalized_relative = relative.replace("\\", "/")
    target = (root / relative).resolve()
    if not target.is_relative_to(root):
        issues.append(f"{label}: 路径超出允许目录")
        return
    if (
        not normalized_relative.startswith("questions/")
        or not target.is_relative_to((root / "questions").resolve())
    ):
        issues.append(f"{label}: 图片路径必须位于 questions/ 目录")
        return
    if not target.is_file():
        issues.append(f"{label}: 文件不存在 {relative}")
        return
    if target.suffix.lower() not in IMAGE_SUFFIXES:
        issues.append(f"{label}: 图片扩展名必须是 .png、.jpg、.jpeg 或 .webp")
        return
    if target.stat().st_size > MAX_IMAGE_BYTES:
        issues.append(f"{label}: 图片不能超过 10 MiB")
        return
    if not image_signature_matches(target):
        issues.append(f"{label}: 图片内容与扩展名不匹配")
        return
    public_url = f"{settings.MEDIA_URL.rstrip('/')}/{normalized_relative}"
    if public_url not in rendered_image_sources(markdown_sources):
        issues.append(f"{label}: 图片未在题干、答案或解析中引用")


class Command(BaseCommand):
    help = "Validate and import non-mathematics A papers and questions from XLSX files."

    def add_arguments(self, parser):
        parser.add_argument("--inventory")
        parser.add_argument("--questions")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--create-templates", dest="create_templates")

    def handle(self, *args, **options):
        if options["create_templates"]:
            self.write_templates(Path(options["create_templates"]))
            return
        if not options["inventory"] or not options["questions"]:
            raise CommandError("必须同时提供 --inventory 和 --questions。")

        paper_rows = self.read_rows(Path(options["inventory"]), INVENTORY_HEADERS, "inventory")
        question_rows = self.read_rows(Path(options["questions"]), QUESTION_HEADERS, "questions")
        issues = []
        if not paper_rows:
            issues.append("inventory: 至少需要一行数据")
        if not question_rows:
            issues.append("questions: 至少需要一行数据")
        papers = self.validate_papers(paper_rows, issues)
        questions = self.validate_questions(question_rows, papers, issues)
        self.validate_counts(papers, questions, issues)

        formula_items = []
        for row in questions:
            label = f"questions 第{row['_line']}行"
            formula_items.extend([
                (f"{label} stem_md", str(row.get("stem_md") or "")),
                (f"{label} answer_md", str(row.get("answer_md") or "")),
                (f"{label} solution_md", str(row.get("solution_md") or "")),
            ])
        issues.extend(validate_markdown_formulas(formula_items))

        if issues:
            raise CommandError("导入校验失败:\n" + "\n".join(f"- {issue}" for issue in issues))
        if options["dry_run"]:
            self.stdout.write(self.style.SUCCESS(
                f"Dry-run passed: {len(papers)} papers, {len(questions)} questions."
            ))
            return

        with transaction.atomic():
            self.upsert(papers, questions)
        self.stdout.write(self.style.SUCCESS(
            f"Imported {len(papers)} papers and {len(questions)} questions."
        ))

    def write_templates(self, directory):
        directory.mkdir(parents=True, exist_ok=True)
        for filename, headers in (
            ("source_inventory.xlsx", INVENTORY_HEADERS),
            ("questions.xlsx", QUESTION_HEADERS),
        ):
            workbook = Workbook()
            workbook.active.append(headers)
            workbook.save(directory / filename)
        self.stdout.write(self.style.SUCCESS(f"Templates created in {directory}"))

    def read_rows(self, path, required_headers, label):
        if not path.is_file():
            raise CommandError(f"{label}: 文件不存在 {path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = workbook.active.iter_rows(values_only=True)
            try:
                first_row = next(rows)
            except StopIteration as error:
                raise CommandError(f"{label}: 工作簿为空") from error
            headers = [str(value or "").strip() for value in first_row]
            missing = [header for header in required_headers if header not in headers]
            if missing:
                raise CommandError(f"{label}: 缺少列 {', '.join(missing)}")
            output = []
            for line, values in enumerate(rows, start=2):
                row = dict(zip(headers, values))
                if any(value not in (None, "") for value in row.values()):
                    row["_line"] = line
                    output.append(row)
            return output
        finally:
            workbook.close()

    def validate_papers(self, rows, issues):
        papers = {}
        valid_stages = {value for value, _ in Paper.Stage.choices}
        for row in rows:
            label = f"inventory 第{row['_line']}行"
            edition = as_int(row.get("edition"), f"{label} edition", issues)
            stage = str(row.get("stage") or "").strip()
            key = (edition, stage)
            if not 1 <= edition <= 17:
                issues.append(f"{label}: edition 必须在 1—17")
            if stage not in valid_stages:
                issues.append(f"{label}: stage 必须是 preliminary 或 final")
            if key in papers:
                issues.append(f"{label}: 重复试卷 {key}")
            if not str(row.get("original_category_label") or "").strip():
                issues.append(f"{label}: original_category_label 不能为空")
            if not str(row.get("title") or "").strip():
                issues.append(f"{label}: title 不能为空")
            row["_edition"] = edition
            row["_stage"] = stage
            exam_year = row.get("exam_year")
            row["_exam_year"] = (
                None
                if exam_year in (None, "")
                else as_int(exam_year, f"{label} exam_year", issues)
            )
            question_count = as_int(
                row.get("question_count"), f"{label} question_count", issues
            )
            if question_count < 1:
                issues.append(f"{label}: question_count 必须大于 0")
            row["_question_count"] = question_count
            safe_file(settings.MEDIA_ROOT, row.get("pdf_file"), f"{label} pdf_file", issues, pdf=True)
            papers[key] = row
        return papers

    def validate_questions(self, rows, papers, issues):
        valid_types = {value for value, _ in Question.Type.choices}
        knowledge = {point.slug: point for point in KnowledgePoint.objects.all()}
        users = {user.username: user for user in get_user_model().objects.all()}
        seen = set()

        for row in rows:
            label = f"questions 第{row['_line']}行"
            edition = as_int(row.get("edition"), f"{label} edition", issues)
            stage = str(row.get("stage") or "").strip()
            number = str(row.get("question_no") or "").strip()
            key = (edition, stage, number)
            if (edition, stage) not in papers:
                issues.append(f"{label}: 找不到对应 inventory 试卷")
            if not number:
                issues.append(f"{label}: question_no 不能为空")
            if key in seen:
                issues.append(f"{label}: 重复题号 {key}")
            seen.add(key)
            question_type = str(row.get("question_type") or "").strip()
            if question_type not in valid_types:
                issues.append(f"{label}: 非法 question_type {question_type}")
            if not str(row.get("stem_md") or "").strip():
                issues.append(f"{label}: stem_md 不能为空")
            if not str(row.get("solution_md") or "").strip():
                issues.append(f"{label}: solution_md 不能为空")

            primary = str(row.get("primary_knowledge") or "").strip()
            secondary = split_values(row.get("secondary_knowledge"))
            for slug, count in Counter(secondary).items():
                if count > 1:
                    issues.append(f"{label}: 次知识点重复 {slug}")
            for slug in [primary, *secondary]:
                if slug not in knowledge:
                    issues.append(f"{label}: 未知知识点 {slug}")
            if primary in secondary:
                issues.append(f"{label}: 主知识点不能重复出现在次知识点")

            reviewer_name = str(row.get("reviewer") or "").strip()
            if reviewer_name and reviewer_name not in users:
                issues.append(f"{label}: reviewer 用户不存在 {reviewer_name}")
            markdown_sources = tuple(
                str(row.get(field) or "")
                for field in ("stem_md", "answer_md", "solution_md")
            )
            for image in split_values(row.get("image_files")):
                validate_question_image(
                    settings.MEDIA_ROOT,
                    image,
                    f"{label} image_files",
                    markdown_sources,
                    issues,
                )
            safe_file(settings.REVIEW_ROOT, row.get("source_crop"), f"{label} source_crop", issues)

            sort_order = as_int(row.get("sort_order"), f"{label} sort_order", issues)
            source_page = as_int(row.get("source_page"), f"{label} source_page", issues)
            unresolved = as_int(
                row.get("unresolved_ocr_items"), f"{label} unresolved_ocr_items", issues
            )
            katex_errors = as_int(row.get("katex_errors"), f"{label} katex_errors", issues)
            if sort_order < 1:
                issues.append(f"{label}: sort_order 必须大于 0")
            if source_page < 1:
                issues.append(f"{label}: source_page 必须大于 0")
            if unresolved < 0 or katex_errors < 0:
                issues.append(f"{label}: OCR 和 KaTeX 错误数不能为负数")

            text_checked = as_bool(
                row.get("text_checked"), f"{label} text_checked", issues
            )
            confidence = as_ocr_confidence(
                row.get("ocr_confidence"), f"{label} ocr_confidence", issues
            )
            if (
                confidence is not None
                and confidence < MIN_OCR_CONFIDENCE
                and not text_checked
                and unresolved == 0
            ):
                issues.append(
                    f"{label}: OCR 置信度低于 0.90 时必须登记 unresolved_ocr_items"
                )
            validate_ocr_characters(row, label, issues)

            row.update({
                "_edition": edition,
                "_stage": stage,
                "_number": number,
                "_sort_order": sort_order,
                "_source_page": source_page,
                "_unresolved": unresolved,
                "_katex_errors": katex_errors,
                "_ocr_confidence": confidence,
                "_primary": primary,
                "_secondary": secondary,
                "_reviewer": users.get(reviewer_name),
                "_text_checked": text_checked,
                "_formula_checked": as_bool(
                    row.get("formula_checked"), f"{label} formula_checked", issues
                ),
                "_solution_checked": as_bool(
                    row.get("solution_checked"), f"{label} solution_checked", issues
                ),
            })
        return rows

    def validate_counts(self, papers, questions, issues):
        actual = Counter((row["_edition"], row["_stage"]) for row in questions)
        for key, paper in papers.items():
            if actual[key] != paper["_question_count"]:
                issues.append(
                    f"inventory 第{paper['_line']}行: question_count={paper['_question_count']}，实际={actual[key]}"
                )

    def upsert(self, papers, questions):
        paper_objects = {}
        for key, row in papers.items():
            paper, _ = Paper.objects.update_or_create(
                edition=row["_edition"],
                stage=row["_stage"],
                defaults={
                    "original_category_label": str(row["original_category_label"]).strip(),
                    "title": str(row["title"]).strip(),
                    "exam_year": row["_exam_year"],
                    "source_url": str(row.get("source_url") or "").strip(),
                    "pdf_file": str(row.get("pdf_file") or "").strip(),
                    "status": Paper.Status.REVIEWED,
                },
            )
            paper.pdf_file.close()
            paper_objects[key] = paper

        knowledge = {point.slug: point for point in KnowledgePoint.objects.all()}
        active_numbers = {key: set() for key in papers}
        for row in questions:
            paper_key = (row["_edition"], row["_stage"])
            active_numbers[paper_key].add(row["_number"])
            reviewed = all((
                row["_text_checked"], row["_formula_checked"], row["_solution_checked"],
                row["_reviewer"] is not None, row["_unresolved"] == 0, row["_katex_errors"] == 0,
            ))
            question, _ = Question.objects.update_or_create(
                paper=paper_objects[paper_key],
                question_no=row["_number"],
                defaults={
                    "sort_order": row["_sort_order"],
                    "question_type": str(row["question_type"]).strip(),
                    "score": row.get("score") or None,
                    "stem_md": str(row["stem_md"]).strip(),
                    "answer_md": str(row.get("answer_md") or "").strip(),
                    "solution_md": str(row["solution_md"]).strip(),
                    "source_page": row["_source_page"],
                    "source_crop": str(row.get("source_crop") or "").strip(),
                    "text_checked": row["_text_checked"],
                    "formula_checked": row["_formula_checked"],
                    "solution_checked": row["_solution_checked"],
                    "unresolved_ocr_items": row["_unresolved"],
                    "katex_errors": row["_katex_errors"],
                    "reviewed_by": row["_reviewer"],
                    "reviewed_at": timezone.now() if reviewed else None,
                    "status": Question.Status.REVIEWED if reviewed else Question.Status.DRAFT,
                },
            )
            QuestionKnowledgePoint.objects.filter(question=question).delete()
            QuestionKnowledgePoint.objects.create(
                question=question,
                knowledge_point=knowledge[row["_primary"]],
                is_primary=True,
            )
            for slug in row["_secondary"]:
                QuestionKnowledgePoint.objects.create(
                    question=question,
                    knowledge_point=knowledge[slug],
                    is_primary=False,
                )

        stale_updated_at = timezone.now()
        for key, paper in paper_objects.items():
            Question.objects.filter(paper=paper).exclude(
                question_no__in=active_numbers[key]
            ).update(status=Question.Status.DRAFT, updated_at=stale_updated_at)
